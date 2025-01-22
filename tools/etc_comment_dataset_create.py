import os
import shutil
from torch.utils.data import Dataset
import openpyxl
    
def main():
    workbook = openpyxl.load_workbook("./data/评语data+分数 2024.9.4.xlsx")
    sheet = workbook.active
    
    row_tag = [3,4,5,6,7,8,9,10,
               11,12,13,14,15,16,17,18,19,20,
               21,22,23,24,25,26,27,28,29,30,
               31,32,33,34,35,36,37,38,39,40]
    column_tag = [22,23]
    
    for i in range(38):
        for j in range(2):
            cell_value = sheet.cell(row=row_tag[i], column=column_tag[j]).value
            if cell_value == True:
                destination_folder = os.path.join("results", sheet.cell(row=2, column=column_tag[j]).value)
                os.makedirs(destination_folder, exist_ok=True)
                source_folder = os.path.join("data", sheet.cell(row=row_tag[i], column=1).value)
                for filename in os.listdir(source_folder):
                    source_path = os.path.join(source_folder, filename)
                    destination_path = os.path.join(destination_folder, filename)
                    shutil.copy2(source_path, destination_path)

                
    
if __name__ == "__main__":
    main()